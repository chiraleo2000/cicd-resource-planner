# -*- coding: utf-8 -*-
"""สร้างไฟล์ Excel ตารางเครื่องมือ CI/CD + เครื่องคำนวณ Co-location (มีสูตรจริง)

รัน:  python3 scripts/build_xlsx.py [output.xlsx]
"""
from __future__ import annotations
import os, sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import catalog_data as C            # noqa: E402
import engine as E                  # noqa: E402
from build_catalog import enrich_tools  # noqa: E402
from openpyxl import Workbook       # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter                            # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation            # noqa: E402
from openpyxl.comments import Comment                                   # noqa: E402

FONT = "Tahoma"          # ฟอนต์มาตรฐานที่แสดงภาษาไทยได้ถูกต้องทั้งบน Excel Windows/Mac
NAVY = "1F3864"
ORANGE = "ED7D31"
GREY = "F2F2F2"
BLUE_IN = "0000FF"       # ตัวเลขที่ผู้ใช้กรอกเอง
GREEN_LINK = "008000"    # อ้างอิงข้ามชีท
YELLOW = "FFFF00"
RED = "FF0000"

STAGE_COLORS = {1: "DDEBF7", 2: "FFF2CC", 3: "FCE4D6", 4: "E2EFDA", 5: "E4DFEC", 6: "FBE5D6"}

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ชื่อชีท
S_README = "00_อ่านก่อน"
S_TOOLS = "01_Tool_Catalog"
S_FREQ = "02_Freq_Weight"
S_CALC = "03_Colocation_Calc"
S_VM = "04_VM_Sizing"
S_PARAM = "05_Params"
S_STORE = "06_Storage_LongTerm"
S_COMP = "07_Compliance_Matrix"
S_CAPMTX = "08_Tool_Capability"
S_PROFILE = "09_Profile_Comparison"
S_ARCH = "10_Archetypes"
S_FW = "11_Frameworks"

TOOLS = enrich_tools()   # เครื่องมือ + การจับคู่ compliance (ไทย / สากล)
NT = len(TOOLS)
R0 = 3                      # แถวแรกของข้อมูลใน 03
R1 = R0 + NT - 1            # แถวสุดท้าย
NVM = 10                    # จำนวนคอลัมน์ VM ที่รองรับ

# ผังเครื่องเริ่มต้นในชีท 03/04 = ผังอ้างอิง 6 เครื่องสำหรับภาครัฐ (ครบตามมาตรฐานบังคับ)
# VM ที่เหลือเว้นว่างให้กรอกเอง — ผังแบบอื่นดูได้ที่ชีท 10_Archetypes
DEFAULT_ARCH = next(a for a in C.ARCHETYPES if a["id"] == "arch-6vm-gov")
PRESET_VMS = [dict(host=vm["host"], role=vm["role_th"], tools=set(vm["tools"]), spec=None)
              for vm in DEFAULT_ARCH["vms"]]


def hdr(ws, row, values, fill=NAVY, color="FFFFFF", size=9, wrap=True, height=42):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=size, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = BORDER
    ws.row_dimensions[row].height = height


def title(ws, text, sub=None, ncols=8):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    if sub:
        ws.cell(row=1, column=3, value=sub).font = Font(name=FONT, size=9, italic=True, color="808080")
    ws.row_dimensions[1].height = 24


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def cellfont(c, bold=False, size=9, color="000000", italic=False):
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    return c


# =========================================================================== #
def sheet_params(wb):
    ws = wb.create_sheet(S_PARAM)
    title(ws, "05 — พารามิเตอร์ของโมเดล (แก้ไขได้)",
          "ช่องตัวอักษรสีน้ำเงิน = ผู้ใช้แก้ไขได้ / สีดำ = คำนวณอัตโนมัติ")
    widths(ws, {"A": 34, "B": 14, "C": 62, "E": 12, "F": 12, "G": 12})

    rows = [
        ("โหมดการคำนวณ (mode)", "strict",
         "strict = ใช้เงื่อนไข B1 บวกทุกเครื่องมือถ่วงน้ำหนัก (ตามโจทย์ที่กำหนด, ปลอดภัยที่สุด) | "
         "realistic = ใช้ B2 บวกข้ามกลุ่มและใช้ค่าสูงสุดในกลุ่มที่รันเรียงกัน"),
        ("Scale Factor ปริมาณข้อมูล", 1.0,
         "1.0 = ค่าฐาน UAT/Production ขนาดเล็ก (~10 builds/วัน, 1-3 แอป, ทีม 5-15 คน). "
         "ทีม 50 คน / 50 builds ต่อวัน ให้ใช้ 3-5"),
        ("OS Reserve — vCPU", C.MODEL["os_reserve_vcpu"], "กันไว้ให้ OS + Container Runtime ของทุก VM"),
        ("OS Reserve — RAM (GB)", C.MODEL["os_reserve_ram_gb"], "กันไว้ให้ OS + Docker/containerd + Agent ต่าง ๆ"),
        ("OS Reserve — Disk (GB)", C.MODEL["os_reserve_disk_gb"], "ระบบปฏิบัติการ Ubuntu + swap + log ระบบ"),
        ("Disk Free Ratio", C.MODEL["disk_free_ratio"],
         "ต้องเหลือพื้นที่ว่างเป็นสัดส่วนนี้เสมอ (0.25 = เหลือว่าง 25%) เพราะ Docker/Elasticsearch จะหยุดทำงานเมื่อ disk ใกล้เต็ม"),
        ("w_base (น้ำหนักต่ำสุด)", C.MODEL["w_base"], "คือ 20% — ใช้กับเครื่องมือที่รันตามคำสั่ง/รอบ Release"),
        ("w_span (ช่วงน้ำหนัก)", C.MODEL["w_span"], "คือ 0.40 ทำให้น้ำหนักเดี่ยวสูงสุด = 0.20 + 0.40 = 0.60 (60%) เมื่ออยู่เครื่องเดียว"),
    ]
    r = 3
    for label, val, note in rows:
        cellfont(ws.cell(row=r, column=1, value=label), bold=True)
        c = ws.cell(row=r, column=2, value=val)
        cellfont(c, color=BLUE_IN, bold=True)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = BORDER
        cellfont(ws.cell(row=r, column=3, value=note), size=8, color="595959").alignment = \
            Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    dv = DataValidation(type="list", formula1='"strict,realistic"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    ws["B8"].number_format = "0.00"

    # horizons  -> B13, B14, B15, B16
    cellfont(ws.cell(row=12, column=1, value="ช่วงเวลาประเมินผลลัพธ์ระยะยาว (เดือน)"), bold=True)
    for i, h in enumerate(C.MODEL["horizons"]):
        rr = 13 + i
        cellfont(ws.cell(row=rr, column=1, value=f"H{i+1}"), bold=True)
        c = ws.cell(row=rr, column=2, value=h)
        cellfont(c, color=BLUE_IN, bold=True)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = BORDER
        cellfont(ws.cell(row=rr, column=3,
                        value=("ค่านี้ใช้ในคอลัมน์ 'Data @H3' ของชีท 04_VM_Sizing" if i == 2 else "")),
                 size=8, italic=True, color="808080")

    # ladders (คอลัมน์ E-G)
    cellfont(ws.cell(row=2, column=5, value="Allocation Ladder — ปัดขึ้นเสมอ"), bold=True, size=10,
             color=NAVY)
    for col, lab in ((5, "vCPU"), (6, "RAM (GB)"), (7, "Disk (GB)")):
        c = ws.cell(row=3, column=col, value=lab)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    maxlen = max(len(C.MODEL["vcpu_ladder"]), len(C.MODEL["ram_ladder"]), len(C.MODEL["disk_ladder"]))
    assert maxlen <= 20, "ladder ยาวเกินช่วง E4:G23 ที่สูตรอ้างถึง"
    for i in range(maxlen):
        rr = 4 + i
        for col, arr in ((5, C.MODEL["vcpu_ladder"]), (6, C.MODEL["ram_ladder"]),
                         (7, C.MODEL["disk_ladder"])):
            if i < len(arr):
                c = ws.cell(row=rr, column=col, value=arr[i])
                cellfont(c, size=8)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")
    return ws


P = lambda ref: f"'{S_PARAM}'!{ref}"   # noqa: E731
MODE = P("$B$3")
SCALE = P("$B$4")
OSV = P("$B$5")
OSR = P("$B$6")
OSD = P("$B$7")
FREE = P("$B$8")
WBASE = P("$B$9")
WSPAN = P("$B$10")
H_CELLS = [P(f"$B${13 + i}") for i in range(4)]
LAD_V = P("$E$4:$E$23")
LAD_R = P("$F$4:$F$23")
LAD_D = P("$G$4:$G$23")


def ladder_formula(value_ref, ladder_ref):
    """คืนค่าขั้นการจัดสรรที่เล็กที่สุดซึ่ง >= value (ปัดขึ้น)"""
    return (f"=IFERROR(INDEX({ladder_ref},COUNTIF({ladder_ref},\"<\"&{value_ref})+1),"
            f"CEILING({value_ref},64))")


# =========================================================================== #
def sheet_freq(wb):
    ws = wb.create_sheet(S_FREQ)
    title(ws, "02 — ชั้นความถี่การรัน และการแปลงเป็นน้ำหนัก (Duty Weight)",
          "w_solo = w_base + w_span x activity_index  ->  ช่วงเดี่ยว 20% - 60%; บน VM ใช้ w_max(n) ladder")
    widths(ws, {"A": 14, "B": 44, "C": 14, "D": 14, "E": 12, "F": 70})
    hdr(ws, 2, ["freq_id", "ความถี่การรัน", "ครั้ง/วัน", "activity_index",
                "น้ำหนัก w", "เหตุผล / ข้อควรระวัง"])
    for i, f in enumerate(C.FREQ_CLASSES):
        r = 3 + i
        vals = [f["id"], f["label_th"], f["runs_per_day"], f["activity_index"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="center" if j != 2 else "left")
        c = ws.cell(row=r, column=5, value=f"={WBASE}+{WSPAN}*D{r}")
        cellfont(c, bold=True, color=GREEN_LINK)
        c.number_format = "0%"
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=6, value=f["note_th"])
        cellfont(c, size=8, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.row_dimensions[r].height = 34
    ws["A11"] = ("ช่วง 20-60%: เครื่องมือตามคำสั่งบวกที่ 20% (กันที่ไว้ให้รันได้) "
                 "เครื่องมือ resident บนเครื่องเดียวบวกที่ 60% แล้วลดตามจำนวนเครื่องมือร่วมเครื่อง "
                 "ลงถึงพื้น 20% เมื่อมี 8 ตัวขึ้นไป")
    cellfont(ws["A11"], size=8, italic=True, color="808080")
    ws.merge_cells("A11:F11")
    ws.row_dimensions[11].height = 30
    return ws


# =========================================================================== #
TOOL_COLS = [
    ("A", 20, "tool_id"), ("B", 46, "เครื่องมือ"), ("C", 7, "Stage"), ("D", 26, "หมวด (ขั้นตอนย่อย)"),
    ("E", 11, "Core/Opt"), ("F", 9, "Grade"), ("G", 20, "License"),
    ("H", 11, "min vCPU"), ("I", 11, "min RAM (GB)"), ("J", 12, "min Disk OS (GB)"),
    ("K", 11, "rec vCPU"), ("L", 11, "rec RAM (GB)"), ("M", 12, "rec Disk OS (GB)"),
    ("N", 10, "Resident 24/7"), ("O", 12, "Idle RAM (GB)"), ("P", 12, "Conc Group"),
    ("Q", 13, "freq_id"), ("R", 10, "น้ำหนัก w"),
    ("S", 12, "Data GB/วัน"), ("T", 12, "Retention (วัน)"), ("U", 11, "Index OH"), ("V", 11, "Growth/ปี"),
    ("W", 34, "Capabilities ที่ตอบได้"),
    ("X", 44, "มาตรฐานไทยที่ช่วยตอบ"), ("Y", 34, "มาตรฐานสากลที่ช่วยตอบ"),
    ("Z", 32, "มาตรการที่ตอบได้ครบด้วยตัวเอง"), ("AA", 44, "มาตรการที่ช่วยตอบบางส่วน"),
    ("AB", 10, "จำนวน\nมาตรการ"),
    ("AC", 46, "ทางเลือก Enterprise-Grade"), ("AD", 40, "ทางเลือก Open Source อื่น"),
    ("AE", 72, "ที่มาของตัวเลข Minimum (Sizing Reference)"),
    ("AF", 78, "ข้อสังเกต / ข้อควรระวังในการวางเครื่อง"),
]


def sheet_tools(wb):
    ws = wb.create_sheet(S_TOOLS)
    title(ws, "01 — ตารางเครื่องมือ CI/CD ทุกประเภท + Resource Requirements (Minimum)",
          f"{NT} เครื่องมือ ครบ 6 Stage ตาม CI/CD Service Blueprint V0.2")
    widths(ws, {c: w for c, w, _ in TOOL_COLS})
    hdr(ws, 2, [h for _, _, h in TOOL_COLS], height=48)
    ws.freeze_panes = "C3"

    for i, t in enumerate(TOOLS):
        r = 3 + i
        fill = STAGE_COLORS[t["stage"]]
        vals = [
            t["id"], t["name"], t["stage"], t["category"], t["core"], t["grade"], t["license"],
            t["min"]["vcpu"], t["min"]["ram_gb"], t["min"]["disk_os_gb"],
            t["rec"]["vcpu"], t["rec"]["ram_gb"], t["rec"]["disk_os_gb"],
            1 if t["resident"] else 0, t["idle_ram_gb"], t["conc_group"], t["freq"], None,
            t["storage"]["data_daily_gb"], t["storage"]["retention_days"],
            t["storage"]["index_overhead"], t["storage"]["growth_yr"],
            ", ".join(t["capabilities"]),
            t["compliance"]["frameworks_th_text"] or "-",
            t["compliance"]["frameworks_intl_text"] or "-",
            ", ".join(t["compliance"]["controls_full"]) or "-",
            ", ".join(t["compliance"]["controls_partial"]) or "-",
            t["compliance"]["control_count"],
            "; ".join(t["enterprise_alt"]), "; ".join(t["oss_alt"]),
            t["sizing_ref"], t["note_th"],
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c, size=8)
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=(j in (2, 4, 23, 24, 25, 26, 27, 29, 30, 31, 32)),
                                    vertical="top",
                                    horizontal="center" if (3 <= j <= 22 and j != 4) or j == 28
                                    else "left")
        # น้ำหนัก w = ดึงจากชีท 02 ตาม freq_id
        c = ws.cell(row=r, column=18,
                    value=f"=INDEX('{S_FREQ}'!$E$3:$E$9,MATCH(Q{r},'{S_FREQ}'!$A$3:$A$9,0))")
        cellfont(c, size=8, bold=True, color=GREEN_LINK)
        c.number_format = "0%"
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        for col in (21, 22):
            ws.cell(row=r, column=col).number_format = "0%"
        ws.row_dimensions[r].height = 58

    n = 3 + NT
    ws.cell(row=n + 1, column=1, value="สรุป")
    cellfont(ws.cell(row=n + 1, column=1), bold=True)
    for col, lab in ((8, "H"), (9, "I"), (10, "J")):
        c = ws.cell(row=n + 1, column=col, value=f"=SUM({lab}3:{lab}{R1})")
        cellfont(c, bold=True)
        c.border = BORDER
    ws.cell(row=n + 2, column=1,
            value="ผลรวมข้างบนคือ 'ถ้าแยกทุกเครื่องมือคนละเครื่อง' ซึ่งไม่ใช่วิธีที่ใช้จริง — "
                  "ดูการคำนวณกรณีแชร์เครื่องที่ชีท 03 และ 04")
    cellfont(ws.cell(row=n + 2, column=1), size=8, italic=True, color=RED)

    ws.cell(row=n + 4, column=1, value="หมายเหตุที่มาของข้อมูล").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    notes = [
        "โครงสร้าง 6 Stage และรายชื่อเครื่องมือ: CI/CD Service Blueprint V0.2",
        "ข้อกำหนด compliance: แนวปฏิบัติการพัฒนาซอฟต์แวร์ กฎระเบียบเกี่ยวข้องทางไซเบอร์และสถาปัตยกรรมระบบที่มั่นคงปลอดภัย V0.2",
        "ค่า min/rec: เอกสารติดตั้งของผู้พัฒนาแต่ละเครื่องมือ (ดูคอลัมน์ Sizing Reference) "
        "และค่าที่พบจากการใช้งานจริงระดับ UAT/Production ขนาดเล็ก",
        f"ค่าฐานปริมาณข้อมูล: {C.STORAGE_BASELINE_TH}",
        "คอลัมน์ 'มาตรฐานไทย/สากลที่ช่วยตอบ' คำนวณจาก capability ของเครื่องมือเทียบกับข้อกำหนดในชีท 07 "
        "จึงเปลี่ยนตามอัตโนมัติเมื่อเพิ่มข้อกำหนดใหม่",
        "'ตอบได้ครบด้วยตัวเอง' = capability ทุกตัวที่ข้อกำหนดนั้นต้องการ มีอยู่ในเครื่องมือนี้ตัวเดียว | "
        "'ช่วยตอบบางส่วน' = ต้องมีเครื่องมืออื่นประกอบด้วย",
        "การมีเครื่องมือครบ ไม่ได้แปลว่าผ่านมาตรฐาน — ยังต้องตั้งค่าให้ถูกต้องและมีหลักฐานการตรวจสอบตามที่กฎหมายกำหนด",
        "ตัวเลขทั้งหมดเป็น 'ค่าตั้งต้นสำหรับประเมิน' ต้องปรับตามผลวัดจริง (baseline measurement) หลังติดตั้ง 2-4 สัปดาห์",
    ]
    for k, txt in enumerate(notes):
        c = ws.cell(row=n + 5 + k, column=1, value="• " + txt)
        cellfont(c, size=8, color="595959")
    return ws


# =========================================================================== #
def sheet_calc(wb):
    ws = wb.create_sheet(S_CALC)
    title(ws, "03 — เครื่องคำนวณกรณีเครื่องเดียวแชร์หลายเครื่องมือ (Co-location Input)",
          "ใส่ 1 ในคอลัมน์ VM ที่ต้องการติดตั้งเครื่องมือนั้น (0 หรือว่าง = ไม่ติดตั้ง) แล้วดูผลที่ชีท 04 · "
          "ค่าเริ่มต้นคือผังอ้างอิง 6 เครื่องสำหรับภาครัฐ — ผังแบบอื่นดูที่ชีท 10")
    cols = {"A": 20, "B": 44, "C": 6, "D": 11, "E": 12, "F": 9, "G": 9, "H": 10, "I": 10,
            "J": 9, "K": 10, "L": 11, "M": 11, "N": 11, "O": 10}
    for i in range(4):
        cols[get_column_letter(16 + i)] = 12
    for i in range(NVM):
        cols[get_column_letter(20 + i)] = 11
    widths(ws, cols)

    heads = ["tool_id", "เครื่องมือ", "Stage", "conc_group", "freq_id", "min vCPU", "min RAM",
             "Idle RAM", "Resident", "Install GB", "GB/วัน", "Retention", "Index OH",
             "Growth/ปี", "น้ำหนัก w"]
    heads += [f"Data GB @H{i+1}" for i in range(4)]
    heads += [f"VM{i+1:02d}" for i in range(NVM)]
    hdr(ws, 2, heads, height=46)
    ws.freeze_panes = "C3"

    for i, t in enumerate(TOOLS):
        r = R0 + i
        fill = STAGE_COLORS[t["stage"]]
        base = [t["id"], t["name"], t["stage"], t["conc_group"], t["freq"],
                t["min"]["vcpu"], t["min"]["ram_gb"], t["idle_ram_gb"],
                1 if t["resident"] else 0, t["storage"]["install_gb"],
                t["storage"]["data_daily_gb"], t["storage"]["retention_days"],
                t["storage"]["index_overhead"], t["storage"]["growth_yr"]]
        for j, v in enumerate(base, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c, size=8)
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j == 2 else "center", vertical="center")
        # w
        c = ws.cell(row=r, column=15,
                    value=f"=INDEX('{S_FREQ}'!$E$3:$E$9,MATCH(E{r},'{S_FREQ}'!$A$3:$A$9,0))")
        cellfont(c, size=8, bold=True, color=GREEN_LINK)
        c.number_format = "0%"
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDER
        # Data GB @ H1..H4
        for k in range(4):
            col = 16 + k
            f = (f"=$K{r}*{SCALE}*POWER(1+$N{r},{H_CELLS[k]}/12)"
                 f"*MIN($L{r},{H_CELLS[k]}*30.44)*(1+$M{r})")
            c = ws.cell(row=r, column=col, value=f)
            cellfont(c, size=8, color=GREEN_LINK)
            c.number_format = "#,##0.0"
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
        # VM selection
        for v in range(NVM):
            col = 20 + v
            val = 0
            if v < len(PRESET_VMS) and t["id"] in PRESET_VMS[v]["tools"]:
                val = 1
            c = ws.cell(row=r, column=col, value=val)
            cellfont(c, size=8, bold=(val == 1), color=BLUE_IN)
            c.fill = PatternFill("solid", fgColor=("D9E1F2" if val else "FFFFFF"))
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22

    # แถวหัวชื่อ VM (row 1) เพื่อให้อ่านง่าย
    for v in range(NVM):
        col = 20 + v
        nm = PRESET_VMS[v]["host"] if v < len(PRESET_VMS) else f"(ว่าง) VM{v+1:02d}"
        c = ws.cell(row=1, column=col, value=nm)
        cellfont(c, size=8, bold=True, color=ORANGE)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    tot = R1 + 1
    cellfont(ws.cell(row=tot, column=1, value="จำนวนเครื่องมือบน VM นี้"), bold=True, size=8)
    for v in range(NVM):
        L = get_column_letter(20 + v)
        c = ws.cell(row=tot, column=20 + v, value=f"=SUM({L}{R0}:{L}{R1})")
        cellfont(c, bold=True, size=9)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    return ws


# =========================================================================== #
VM_COLS = [
    ("A", 24, "VM"), ("B", 40, "บทบาทหน้าที่"), ("C", 8, "จำนวน\nเครื่องมือ"),
    ("D", 10, "A: Peak-Max\nvCPU"), ("E", 10, "A: Peak-Max\nRAM"),
    ("F", 10, "B1 Strict\nvCPU"), ("G", 10, "B1 Strict\nRAM"),
    ("H", 10, "B2 Realistic\nvCPU"), ("I", 10, "B2 Realistic\nRAM"),
    ("J", 10, "C: Resident\nvCPU"), ("K", 10, "C: Resident\nRAM"),
    ("L", 10, "MAX(A,B,C)\nvCPU"), ("M", 10, "MAX(A,B,C)\nRAM"),
    ("N", 10, "+OS\nvCPU"), ("O", 10, "+OS\nRAM"),
    ("P", 11, "จัดสรร\nvCPU"), ("Q", 11, "จัดสรร\nRAM (GB)"),
    ("R", 11, "Install\nรวม (GB)"), ("S", 12, "Data @H3\n(GB)"),
    ("T", 12, "ต้องมี Disk OS\n(GB)"), ("U", 12, "ต้องมี Disk Data\n(GB)"),
    ("V", 12, "จัดสรร\nDisk OS"), ("W", 12, "จัดสรร\nDisk Data"),
    ("X", 10, "spec จริง\nvCPU"), ("Y", 10, "spec จริง\nRAM"),
    ("Z", 11, "spec จริง\nDisk OS"), ("AA", 11, "spec จริง\nDisk Data"),
    ("AB", 9, "ส่วนต่าง\nvCPU"), ("AC", 9, "ส่วนต่าง\nRAM"),
    ("AD", 10, "ส่วนต่าง\nDisk OS"), ("AE", 10, "ส่วนต่าง\nDisk Data"),
    ("AF", 20, "ผลประเมิน"), ("AG", 16, "ตัวกำหนด vCPU"), ("AH", 16, "ตัวกำหนด RAM"),
]


def sheet_vm(wb):
    ws = wb.create_sheet(S_VM)
    title(ws, "04 — ผลการคำนวณทรัพยากรต่อ VM (Minimum ที่ต้องจัดสรร)",
          "สูตรทั้งหมดอ้างอิงชีท 03 และ 05 — เปลี่ยนการติดตั้งที่ชีท 03 แล้วตัวเลขที่นี่จะเปลี่ยนตาม")
    widths(ws, {c: w for c, w, _ in VM_COLS})
    hdr(ws, 2, [h for _, _, h in VM_COLS], height=58)
    ws.freeze_panes = "C3"

    G = f"'{S_CALC}'"
    rng = lambda L: f"{G}!${L}${R0}:${L}${R1}"   # noqa: E731

    for v in range(NVM):
        r = 3 + v
        SEL = f"{G}!{get_column_letter(20+v)}${R0}:{get_column_letter(20+v)}${R1}"
        preset = PRESET_VMS[v] if v < len(PRESET_VMS) else None

        cellfont(ws.cell(row=r, column=1, value=(preset["host"] if preset else f"VM{v+1:02d}")),
                 bold=True, size=9, color=NAVY)
        cellfont(ws.cell(row=r, column=2, value=(preset["role"] if preset else "(กรอกบทบาทหน้าที่)")),
                 size=8).alignment = Alignment(wrap_text=True, vertical="top")

        F = {}
        F["C"] = f"=SUM({SEL})"
        # A : Peak-Max
        F["D"] = f"=SUMPRODUCT(MAX({SEL}*{rng('F')}))"
        F["E"] = f"=SUMPRODUCT(MAX({SEL}*{rng('G')}))"
        # B1 : Weighted sum ทุกตัว
        F["F"] = f"=SUMPRODUCT({SEL},{rng('O')},{rng('F')})"
        F["G"] = f"=SUMPRODUCT({SEL},{rng('O')},{rng('G')})"
        # B2 : resident บวก + กลุ่มอื่นใช้ค่าสูงสุด
        def b2(metric):
            res = f'SUMPRODUCT({SEL},--({rng("D")}="resident"),{rng("O")},{rng(metric)})'
            parts = [res]
            for g in ("ci_seq", "async", "load"):
                parts.append(f'SUMPRODUCT(MAX({SEL}*--({rng("D")}="{g}")*{rng("O")}*{rng(metric)}))')
            return "=" + "+".join(parts)
        F["H"] = b2("F")
        F["I"] = b2("G")
        # C : Resident floor
        F["J"] = f"=SUMPRODUCT({SEL},{rng('I')})*0.25"
        F["K"] = f"=SUMPRODUCT({SEL},{rng('I')},{rng('H')})"
        # MAX(A, B ตามโหมด, C)
        F["L"] = f'=MAX(D{r},IF({MODE}="strict",F{r},H{r}),J{r})'
        F["M"] = f'=MAX(E{r},IF({MODE}="strict",G{r},I{r}),K{r})'
        F["N"] = f"=L{r}+{OSV}"
        F["O"] = f"=M{r}+{OSR}"
        F["P"] = ladder_formula(f"N{r}", LAD_V)
        F["Q"] = ladder_formula(f"O{r}", LAD_R)
        # Storage
        F["R"] = f"=SUMPRODUCT({SEL},{rng('J')})"
        F["S"] = f"=SUMPRODUCT({SEL},{rng('R')})"
        F["T"] = f"=({OSD}+R{r})/(1-{FREE})"
        F["U"] = f"=S{r}/(1-{FREE})"
        F["V"] = ladder_formula(f"T{r}", LAD_D)
        F["W"] = f"=IF(U{r}<=0,0,{ladder_formula(f'U{r}', LAD_D)[1:]})"
        # spec จริง
        for col, key in (("X", "vcpu"), ("Y", "ram_gb"), ("Z", "disk_os_gb"), ("AA", "disk_data_gb")):
            val = (preset["spec"] or {}).get(key) if preset and preset.get("spec") else None
            c = ws.cell(row=r, column=ws[col + "2"].column, value=val)
            cellfont(c, color=BLUE_IN, bold=True, size=9)
            c.fill = PatternFill("solid", fgColor=YELLOW)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        F["AB"] = f"=IF(X{r}=\"\",\"\",X{r}-P{r})"
        F["AC"] = f"=IF(Y{r}=\"\",\"\",Y{r}-Q{r})"
        F["AD"] = f"=IF(Z{r}=\"\",\"\",Z{r}-V{r})"
        F["AE"] = f"=IF(AA{r}=\"\",\"\",AA{r}-W{r})"
        F["AF"] = (f'=IF(X{r}="","(ยังไม่กรอก spec)",'
                   f'IF(OR(AB{r}<0,AC{r}<0),"ไม่พอ: CPU/RAM",'
                   f'IF(OR(AD{r}<0,AE{r}<0),"เสี่ยง: Disk ไม่พอ","พอเพียง")))')
        F["AG"] = (f'=IF(L{r}=D{r},"A Peak-Max",'
                   f'IF(L{r}=J{r},"C Resident","B Weighted"))')
        F["AH"] = (f'=IF(M{r}=E{r},"A Peak-Max",'
                   f'IF(M{r}=K{r},"C Resident","B Weighted"))')

        for col, formula in F.items():
            c = ws.cell(row=r, column=ws[col + "2"].column, value=formula)
            cellfont(c, size=9, bold=col in ("P", "Q", "V", "W", "AF"),
                     color=(NAVY if col in ("P", "Q", "V", "W") else "000000"))
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", wrap_text=(col == "AF"))
            if col in ("D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
                c.number_format = "#,##0.00"
            elif col in ("R", "S", "T", "U"):
                c.number_format = "#,##0"
            elif col in ("P", "Q", "V", "W", "AB", "AC", "AD", "AE"):
                c.number_format = "#,##0"
            if col in ("P", "Q", "V", "W"):
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        ws.row_dimensions[r].height = 32

    # อธิบายวิธีคำนวณ
    r = 3 + NVM + 2
    ws.cell(row=r, column=1, value="วิธีคำนวณ (ตามเงื่อนไขที่กำหนด)").font = \
        Font(name=FONT, size=11, bold=True, color=NAVY)
    explain = [
        ("เงื่อนไข 1 — A: Peak-Max", "คือ MAX( minimum ของทุกเครื่องมือบนเครื่องนั้น )  "
         "ตีความว่า ณ เวลาหนึ่งมีเครื่องมือเดียวที่ทำงานหนักสุด — เป็นพื้นขั้นต่ำที่ต้องมีเพื่อให้ตัวที่หนักสุดรันผ่าน"),
        ("เงื่อนไข 2 — B1: Weighted-Sum (strict)", "คือ ผลรวมของ ( minimum x w ) ทุกเครื่องมือ  "
         "โดย w = 50% ถึง 95% ตามความถี่ที่เครื่องมือต้องรันหรือรันค้างหลังบ้าน (ดูชีท 02) "
         "— เป็นค่าที่โจทย์กำหนด และเป็นค่าปลอดภัยที่สุด"),
        ("เงื่อนไข 2 — B2: Weighted-Sum (realistic)", "คือ ผลรวมกลุ่ม resident (บวกทุกตัว) "
         "+ ค่าสูงสุดของกลุ่ม ci_seq + ค่าสูงสุดของกลุ่ม async + ค่าสูงสุดของกลุ่ม load  "
         "เพราะขั้นตอนภายใน Pipeline รอบเดียวกันทำงานเรียงต่อกัน ไม่ได้รันพร้อมกัน"),
        ("ตัวตรวจ — C: Resident Floor", "คือ ผลรวม Idle RAM ของเครื่องมือที่รันค้าง 24/7  "
         "ถ้า A และ B ต่ำกว่า C แปลว่าเครื่องบูตขึ้นมาก็เต็มแล้ว จึงต้องยกขึ้นเป็น C"),
        ("ผลลัพธ์", "REQUIRED เท่ากับ MAX(A, B, C) + OS Reserve  ->  ปัดขึ้นตาม Allocation Ladder  "
         "(ตามโจทย์ที่ระบุว่า 'ต้องเป็นค่าที่มากสุดสำหรับ minimum เท่านั้น')"),
        ("Disk OS", "คือ (OS Reserve + ผลรวม Install ของทุกเครื่องมือ) / (1 - Disk Free Ratio)"),
        ("Disk Data", "คือ ผลรวม Data GB ณ ช่วงเวลาที่ประเมิน / (1 - Disk Free Ratio)  "
         "โดย Data GB เท่ากับ GB/วัน x Scale x (1+Growth)^(เดือน/12) x MIN(Retention, เดือน x 30.44) x (1 + Index OH)"),
    ]
    for k, (a, b) in enumerate(explain):
        rr = r + 1 + k
        cellfont(ws.cell(row=rr, column=1, value=a), bold=True, size=8, color=ORANGE)
        c = ws.cell(row=rr, column=3, value=b)
        cellfont(c, size=8, color="404040")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=18)
        ws.row_dimensions[rr].height = 30
    return ws


# =========================================================================== #
def sheet_storage(wb):
    ws = wb.create_sheet(S_STORE)
    title(ws, "06 — ผลลัพธ์ระยะยาวของพื้นที่จัดเก็บ (Storage Long-term Projection)",
          "ต่อ 1 instance ของเครื่องมือ — คูณจำนวน instance และรวมตาม VM ที่ชีท 04")
    widths(ws, {"A": 20, "B": 44, "C": 12, "D": 12, "E": 11, "F": 11, "G": 12,
                "H": 13, "I": 13, "J": 13, "K": 13, "L": 14, "M": 60})
    hdr(ws, 2, ["tool_id", "เครื่องมือ", "Install (GB)", "GB/วัน", "Retention\n(วัน)",
                "Index OH", "Growth/ปี",
                "รวม @12 ด.", "รวม @24 ด.", "รวม @36 ด.", "รวม @60 ด.",
                "ต้องจัดสรร @60 ด.", "ข้อควรทำเพื่อคุมการโต"], height=46)
    ws.freeze_panes = "C3"

    ADVICE = {
        "elasticsearch": "ตั้ง ILM Policy: hot 7 วัน -> warm 30 วัน -> delete ที่ 90 วัน (ตามมาตรฐานขั้นต่ำ 2566); "
                         "เปิด best_compression; แยก index ต่อวันเพื่อลบเป็นก้อน",
        "minio": "ตั้ง Object Lifecycle: ลบ artifact ที่ไม่ใช่ release เมื่อครบ 30-90 วัน; "
                 "เปิด Versioning เฉพาะ bucket ที่จำเป็น เพราะ version เก่ากินที่เท่าตัว",
        "harbor": "ตั้ง Tag Retention Policy (เก็บ 10 tag ล่าสุดต่อ repo) + เปิด Garbage Collection รายสัปดาห์",
        "nexus-repository": "แยก blob store ตามชนิด repo; เปิด Cleanup Policy ของ snapshot/proxy cache "
                            "(เช่น ลบ snapshot หลัง 30 วัน เก็บ release ถาวร); ห้ามเปิด Redeploy บน maven-releases",
        "zot": "ตั้ง retention ของ tag ที่ไม่ใช่ release และจำกัดขนาด blob store",
        "docker-buildkit": "ตั้ง cron `docker system prune -af --filter until=336h` และจำกัด BuildKit cache "
                           "ด้วย --max-used-space",
        "jenkins-agent": "ใช้ cleanWs() ทุกท้าย Pipeline และตั้ง Workspace Cleanup Plugin",
        "jenkins-master": "ตั้ง buildDiscarder(logRotator(numToKeepStr:'50', daysToKeepStr:'180'))",
        "postgresql-tools": "ตั้ง partition ตารางที่โตเร็วเป็นรายเดือน + archive ไป MinIO เมื่อเกิน 1 ปี "
                            "(Audit Trail ภาครัฐต้องเก็บ 7 ปีแต่ไม่จำเป็นต้องอยู่ใน DB หลัก)",
        "prometheus": "ตั้ง --storage.tsdb.retention.time=90d และลด scrape_interval ของ metric ที่ไม่สำคัญ; "
                      "ถ้าต้องเก็บนานให้ส่งต่อ VictoriaMetrics/Thanos",
        "wazuh": "ตั้ง ISM Policy บน Wazuh Indexer + ปิด rule ที่สร้าง event มากเกินจำเป็น",
        "owasp-zap": "เก็บเฉพาะไฟล์รายงาน (HTML/JSON) ไม่เก็บ session file (.session) ที่ใหญ่หลาย GB",
        "sonarqube": "ตั้ง Housekeeping: ลบ snapshot ที่เก่ากว่า 30 วัน, เก็บเฉพาะ analysis ของ branch หลัก",
        "velero-restic": "ใช้ incremental + dedup, กำหนด GFS (Grandfather-Father-Son) แทนการเก็บ full ทุกวัน",
        "gitlab-ce": "ตั้ง housekeeping (git gc) อัตโนมัติ + จำกัดขนาด artifact และ LFS",
        "sftp-nfs": "ย้ายไฟล์เก่ากว่า 90 วันไป MinIO แล้วเก็บเฉพาะ pointer",
        "gpu-training": "เก็บเฉพาะ checkpoint ที่ดีที่สุดและ checkpoint สุดท้าย ลบ intermediate ที่เหลือ",
        "mlflow": "ตั้ง artifact retention ต่อ experiment + ย้าย artifact เก่าไป cold storage",
        "falco": "ปรับ rule ให้แคบก่อน production ไม่งั้น event flood ทำให้ Elasticsearch โตเกินประเมิน",
    }

    for i, t in enumerate(TOOLS):
        r = 3 + i
        s = t["storage"]
        fill = STAGE_COLORS[t["stage"]]
        vals = [t["id"], t["name"], s["install_gb"], s["data_daily_gb"], s["retention_days"],
                s["index_overhead"], s["growth_yr"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c, size=8)
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j == 2 else "center")
        for j in (6, 7):
            ws.cell(row=r, column=j).number_format = "0%"
        for k in range(4):
            col = 8 + k
            f = (f"=$C{r}+$D{r}*{SCALE}*POWER(1+$G{r},{H_CELLS[k]}/12)"
                 f"*MIN($E{r},{H_CELLS[k]}*30.44)*(1+$F{r})")
            c = ws.cell(row=r, column=col, value=f)
            cellfont(c, size=8, color=GREEN_LINK)
            c.number_format = "#,##0.0"
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
        c = ws.cell(row=r, column=12, value=ladder_formula(f"K{r}/(1-{FREE})", LAD_D))
        cellfont(c, size=9, bold=True, color=NAVY)
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        adv = ADVICE.get(t["id"], "ปริมาณข้อมูลน้อย ไม่ต้องมีนโยบายเฉพาะ — รวมอยู่ใน Disk OS ได้")
        c = ws.cell(row=r, column=13, value=adv)
        cellfont(c, size=8, color="404040")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.row_dimensions[r].height = 34

    tot = 3 + NT
    cellfont(ws.cell(row=tot, column=2, value="รวมทุกเครื่องมือ (กรณีติดตั้งครบทุกตัว)"), bold=True)
    for col in ("C", "H", "I", "J", "K"):
        c = ws.cell(row=tot, column=ws[col + "2"].column, value=f"=SUM({col}3:{col}{2+NT})")
        cellfont(c, bold=True, size=9, color=RED)
        c.number_format = "#,##0"
        c.border = BORDER
    return ws


# =========================================================================== #
def sheet_capmatrix(wb):
    ws = wb.create_sheet(S_CAPMTX)
    title(ws, "08 — เมทริกซ์เครื่องมือ x Capability",
          "1 = เครื่องมือนี้ตอบ capability นั้นได้ | ใช้เป็นฐานของการตรวจ compliance ในชีท 07")
    caps = list(C.CAPABILITIES.keys())
    widths(ws, {"A": 20, "B": 40})
    for i in range(len(caps)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 5.5
    hdr(ws, 2, ["tool_id", "เครื่องมือ"] + caps, height=110)
    for i, cap in enumerate(caps):
        ws.cell(row=2, column=3 + i).alignment = Alignment(textRotation=90, horizontal="center",
                                                          vertical="bottom")
    ws.freeze_panes = "C3"
    for i, t in enumerate(TOOLS):
        r = 3 + i
        cellfont(ws.cell(row=r, column=1, value=t["id"]), size=8)
        cellfont(ws.cell(row=r, column=2, value=t["name"]), size=8)
        for j, cap in enumerate(caps):
            on = cap in t["capabilities"]
            c = ws.cell(row=r, column=3 + j, value=1 if on else 0)
            cellfont(c, size=7, bold=on, color=("006100" if on else "BFBFBF"))
            c.fill = PatternFill("solid", fgColor=("C6EFCE" if on else "FFFFFF"))
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 16
    # แถวสรุป: capability นี้ถูกครอบคลุมโดยเครื่องมือที่เลือกไว้ในชีท 03 หรือไม่
    r = 3 + NT + 1
    cellfont(ws.cell(row=r, column=2, value="ครอบคลุมโดยเครื่องมือที่เลือกไว้ในชีท 03? (1=ใช่)"),
             bold=True, size=8, color=NAVY)
    selrange = (f"'{S_CALC}'!$T${R0}:$AC${R1}")
    for j in range(len(caps)):
        L = get_column_letter(3 + j)
        f = (f"=IF(SUMPRODUCT(--(MMULT({selrange},TRANSPOSE(COLUMN({selrange})*0+1))>0),"
             f"{L}{3}:{L}{2+NT})>0,1,0)")
        # MMULT/TRANSPOSE ไม่เสถียรเมื่อเขียนจาก openpyxl -> ใช้คอลัมน์ช่วยแทน
        f = f"=IF(SUMPRODUCT($AE${3}:$AE${2+NT},{L}{3}:{L}{2+NT})>0,1,0)"
        c = ws.cell(row=r, column=3 + j, value=f)
        cellfont(c, size=7, bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    # คอลัมน์ช่วย AE: เครื่องมือถูกเลือกใน VM ใดก็ได้
    hcol = 3 + len(caps) + 1
    HL = get_column_letter(hcol)
    cellfont(ws.cell(row=2, column=hcol, value="เลือกใน VM ใดก็ได้"), bold=True, size=7)
    for i in range(NT):
        rr = 3 + i
        f = f"=IF(SUM('{S_CALC}'!$T${R0+i}:$AC${R0+i})>0,1,0)"
        c = ws.cell(row=rr, column=hcol, value=f)
        cellfont(c, size=7)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[HL].width = 10
    # แก้สูตรแถวสรุปให้ชี้คอลัมน์ช่วยจริง
    for j in range(len(caps)):
        L = get_column_letter(3 + j)
        ws.cell(row=r, column=3 + j).value = \
            f"=IF(SUMPRODUCT(${HL}$3:${HL}${2+NT},{L}3:{L}{2+NT})>0,1,0)"
    return ws, caps, HL


# =========================================================================== #
SEV_TH = {"mandatory": "บังคับ", "conditional": "บังคับเมื่อผลกระทบสูง", "recommended": "แนะนำ"}


def sheet_compliance(wb, caps, capmtx_helper):
    """ชีท 07 — มาตรการที่ต้องทำ (Control) พร้อมสถานะและมาตรฐานที่อ้าง"""
    ws = wb.create_sheet(S_COMP)
    title(ws, "07 — มาตรการที่ต้องทำ (Controls) และสถานะตามเครื่องมือที่เลือก",
          "1 แถว = 1 มาตรการ · มาตรฐานหลายฉบับอ้างมาตรการเดียวกันได้ ดูเลขข้อในคอลัมน์ท้าย")
    widths(ws, {"A": 16, "B": 24, "C": 12, "D": 62, "E": 52, "F": 34, "G": 12, "H": 34, "I": 78})
    hdr(ws, 2, ["control_id", "กลุ่มมาตรการ", "ระดับบังคับ", "มาตรการที่ต้องทำ",
                "คำอธิบายเพิ่มเติม", "Capability ที่ต้องมี", "สถานะ",
                "Capability ที่ยังขาด", "มาตรฐานที่อ้างมาตรการนี้ (เลขข้อ)"], height=42)
    ws.freeze_panes = "A3"

    cap_col = {c: get_column_letter(3 + i) for i, c in enumerate(caps)}
    covered_row = 3 + NT + 1
    grp = C.CONTROL_GROUPS

    for i, ctl in enumerate(C.CONTROLS):
        r = 3 + i
        refs = C.framework_refs(ctl["id"])
        reftxt = "\n".join(f'{C.FRAMEWORK_BY_ID[f]["short_th"]} — {cl}' for f, cl in refs.items())
        vals = [ctl["id"], grp[ctl["group"]], SEV_TH[ctl["severity"]], ctl["title_th"],
                ctl.get("detail_th", ""), ", ".join(ctl["caps"])]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c, size=8, bold=(j == 4))
            c.border = BORDER
            c.alignment = Alignment(wrap_text=(j in (4, 5, 6)), vertical="top",
                                    horizontal="center" if j in (1, 3) else "left")
        terms = [f"'{S_CAPMTX}'!{cap_col[x]}${covered_row}" for x in ctl["caps"]]
        f = f'=IF(({"+".join(terms)})>={len(ctl["caps"])},"ผ่าน","ไม่ผ่าน")'
        c = ws.cell(row=r, column=7, value=f)
        cellfont(c, size=8, bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        parts = [f'IF(\'{S_CAPMTX}\'!{cap_col[x]}${covered_row}=0,"{x} ","")' for x in ctl["caps"]]
        c = ws.cell(row=r, column=8, value="=" + "&".join(parts))
        cellfont(c, size=8, color=RED)
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c = ws.cell(row=r, column=9, value=reftxt or "-")
        cellfont(c, size=8, color="404040")
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34

    n = 3 + len(C.CONTROLS)
    cellfont(ws.cell(row=n + 1, column=4, value="คะแนน Compliance (นับทุกมาตรการในระบบ)"),
             bold=True, size=10, color=NAVY)
    c = ws.cell(row=n + 1, column=7, value=f'=COUNTIF(G3:G{n-1},"ผ่าน")/COUNTA(G3:G{n-1})')
    cellfont(c, bold=True, size=12, color=NAVY)
    c.number_format = "0.0%"
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.border = BORDER
    cellfont(ws.cell(row=n + 2, column=4, value="จำนวนมาตรการที่ไม่ผ่าน"), bold=True, size=9)
    c = ws.cell(row=n + 2, column=7, value=f'=COUNTIF(G3:G{n-1},"ไม่ผ่าน")')
    cellfont(c, bold=True, size=11, color=RED)
    c.border = BORDER
    ws.cell(row=n + 4, column=1,
            value="หมายเหตุ: คะแนนในชีทนี้นับทุกมาตรการที่มีในระบบ ไม่ได้กรองตามมาตรฐานที่โครงการเลือก "
                  "— ถ้าต้องการคะแนนตามชุดมาตรฐานที่เลือกจริง ให้ใช้โปรแกรมเว็บ (แท็บ 3)")
    cellfont(ws.cell(row=n + 4, column=1), size=8, italic=True, color="808080")
    return ws


# =========================================================================== #
def sheet_frameworks(wb):
    """ชีท 11 — ทะเบียนมาตรฐานรายฉบับ + เมทริกซ์มาตรฐาน x มาตรการ"""
    ws = wb.create_sheet(S_FW)
    title(ws, "11 — ทะเบียนกฎหมายและมาตรฐานรายฉบับ",
          "แยกไทย/สากล พร้อมเลขข้อที่อ้างแต่ละมาตรการ · ⚠ = ควรตรวจเลขที่ประกาศก่อนอ้างใน TOR")
    ctl_ids = [c["id"] for c in C.CONTROLS]
    widths(ws, {"A": 22, "B": 9, "C": 30, "D": 58, "E": 26, "F": 52, "G": 9})
    for i in range(len(ctl_ids)):
        ws.column_dimensions[get_column_letter(8 + i)].width = 5.6
    hdr(ws, 2, ["framework_id", "ขอบเขต", "กลุ่ม", "ชื่อมาตรฐาน", "หน่วยงานเจ้าของ",
                "ขอบเขตที่บังคับใช้", "จำนวน\nมาตรการ"] + ctl_ids, height=112)
    for i in range(len(ctl_ids)):
        ws.cell(row=2, column=8 + i).alignment = Alignment(textRotation=90,
                                                          horizontal="center", vertical="bottom")
    ws.freeze_panes = "D3"

    fams = sorted(C.FRAMEWORK_FAMILIES.items(), key=lambda kv: kv[1]["order"])
    r = 3
    for fid, fam in fams:
        rows = [f for f in C.FRAMEWORKS if f["family"] == fid]
        if not rows:
            continue
        c = ws.cell(row=r, column=1, value=f'{fam["label_th"]} ({len(rows)} ฉบับ)')
        cellfont(c, bold=True, size=9, color=NAVY)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        for k in range(2, 8 + len(ctl_ids)):
            ws.cell(row=r, column=k).fill = PatternFill("solid", fgColor="D9E1F2")
        r += 1
        for f in rows:
            fill = "FFF2CC" if f["region"] == "th" else "E2EFDA"
            vals = [f["id"] + (" ⚠" if f.get("verify") else ""),
                    "ไทย" if f["region"] == "th" else "สากล",
                    f["family"], f["name_th"], f.get("authority", ""), f.get("scope_th", ""),
                    len(f["controls"])]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                cellfont(c, size=8, bold=(j == 4))
                c.fill = PatternFill("solid", fgColor=fill)
                c.border = BORDER
                c.alignment = Alignment(wrap_text=(j in (4, 6)), vertical="top",
                                        horizontal="center" if j in (2, 7) else "left")
            for k, cid in enumerate(ctl_ids):
                ref = f["controls"].get(cid)
                txt = (ref["clause"] if isinstance(ref, dict) else ref) if ref else None
                c = ws.cell(row=r, column=8 + k, value=("✓" if txt else None))
                if txt:
                    c.comment = Comment(f'{cid}\n{C.CONTROL_BY_ID[cid]["title_th"]}\n\n'
                                        f'{f["short_th"]} — {txt}', "CI/CD Planner")
                cellfont(c, size=8, bold=bool(txt), color=("006100" if txt else "BFBFBF"))
                c.fill = PatternFill("solid", fgColor=("C6EFCE" if txt else "FFFFFF"))
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    cellfont(ws.cell(row=r + 1, column=1,
                     value="เครื่องหมาย ✓ = มาตรฐานฉบับนั้นอ้างมาตรการนั้น (วางเมาส์บนช่องเพื่อดูเลขข้อ) · "
                           "รายละเอียดมาตรการดูที่ชีท 07"),
             size=8, italic=True, color="808080")
    return ws


# =========================================================================== #
def sheet_profile(wb):
    ws = wb.create_sheet(S_PROFILE)
    title(ws, "09 — เปรียบเทียบประเภทโครงการ 5 แบบ",
          "ใช้เลือก preset ตั้งต้นก่อนปรับรายละเอียด")
    widths(ws, {"A": 12, "B": 24, "C": 12, "D": 14, "E": 16, "F": 20, "G": 26,
                "H": 14, "I": 16, "J": 74, "K": 58})
    hdr(ws, 2, ["profile", "ประเภทโครงการ", "ระดับ\nผลกระทบ", "Security", "เกรดเครื่องมือ\nที่แนะนำ",
                "เวลา Pipeline\nต่อรอบ", "ต้นทุน/ปี (บาท)",
                "Log\nRetention", "Audit\nRetention", "ชุดมาตรฐานตั้งต้นของประเภทนี้",
                "ข้อควรระวัง"], height=46)
    for i, p in enumerate(C.PROFILES):
        r = 3 + i
        vals = [p["id"], p["name_th"], p["impact"], p["security"], p["grade_pref"],
                p["automate_th"], p["cost_yr_thb"],
                f'{p["log_retention_days"]} วัน', f'{p["audit_retention_days"]} วัน',
                f'ชุด "{C.PRESET_LABELS.get(p["framework_preset"], p["framework_preset"])}" '
                f'({len(C.FRAMEWORK_PRESETS.get(p["framework_preset"], []))} ฉบับ): '
                + ", ".join(C.FRAMEWORK_BY_ID[f]["short_th"]
                            for f in C.FRAMEWORK_PRESETS.get(p["framework_preset"], [])),
                p["notes_th"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            cellfont(c, size=8, bold=(j == 2))
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if j in (1, 3, 4, 5, 8, 9) else "left")
        ws.row_dimensions[r].height = 60
    return ws


# =========================================================================== #
def sheet_archetypes(wb):
    ws = wb.create_sheet(S_ARCH)
    title(ws, "10 — ผังเครื่องอ้างอิง (Reference Architectures)",
          "ผังตั้งต้นแบบทั่วไป ไม่ผูกกับโครงการใด — คัดลอกรายการเครื่องมือไปติ๊กในชีท 03 ได้")
    widths(ws, {"A": 22, "B": 34, "C": 22, "D": 52, "E": 8, "F": 9, "G": 9, "H": 10, "I": 11, "J": 62})
    hdr(ws, 2, ["ผัง", "คำอธิบายผัง", "เครื่อง (Host)", "บทบาทหน้าที่", "จำนวน\nเครื่องมือ",
                "vCPU\nที่ควรขอ", "RAM\nที่ควรขอ", "Disk OS\nที่ควรขอ", "Disk Data\nที่ควรขอ",
                "รายการ tool_id ที่ติดตั้งบนเครื่องนี้"], height=46)
    ws.freeze_panes = "C3"

    r = 3
    for ai, a in enumerate(C.ARCHETYPES):
        first = r
        tools_all = sorted({t for v in a["vms"] for t in v["tools"]})
        comp = E.compliance_check(tools_all, a["profile"])
        for vm in a["vms"]:
            calc = E.colocate(vm["tools"], horizon_months=36, mode="realistic")
            al = calc["allocated"]
            vals = [a["name_th"] if vm is a["vms"][0] else None,
                    a["network_th"] if vm is a["vms"][0] else None,
                    vm["host"], vm["role_th"], len(vm["tools"]),
                    al["vcpu"], al["ram_gb"], al["disk_os_gb"], al["disk_data_gb"],
                    ", ".join(vm["tools"])]
            fill = "F2F2F2" if ai % 2 == 0 else "FFFFFF"
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                cellfont(c, size=8, bold=(j == 1))
                c.fill = PatternFill("solid", fgColor=fill)
                c.border = BORDER
                c.alignment = Alignment(wrap_text=(j in (1, 2, 4, 10)), vertical="top",
                                        horizontal="center" if 5 <= j <= 9 else "left")
            ws.row_dimensions[r].height = 40
            r += 1
        if r - first > 1:
            for col in (1, 2):
                ws.merge_cells(start_row=first, start_column=col, end_row=r - 1, end_column=col)
        # แถวสรุปของผัง
        cellfont(ws.cell(row=r, column=3, value="รวมทั้งผัง"), bold=True, size=8, color=NAVY)
        cellfont(ws.cell(row=r, column=4,
                         value=f"เครื่องมือไม่ซ้ำ {len(tools_all)} รายการ · "
                               f"คะแนน Compliance ตามประเภทโครงการ {a['profile']} = {comp['score']}%"),
                 size=8, color=NAVY)
        for j, key in ((6, "vcpu"), (7, "ram_gb"), (8, "disk_os_gb"), (9, "disk_data_gb")):
            tot = sum(E.colocate(v["tools"], horizon_months=36,
                                 mode="realistic")["allocated"][key] for v in a["vms"])
            c = ws.cell(row=r, column=j, value=tot)
            cellfont(c, bold=True, size=9, color=NAVY)
            c.fill = PatternFill("solid", fgColor="E2EFDA")
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        r += 2

    cellfont(ws.cell(row=r, column=1, value="ชุดเครื่องมือตามบทบาท (Role Bundle) ที่ใช้ประกอบผังข้างบน"),
             bold=True, size=11, color=NAVY)
    r += 1
    hdr(ws, r, ["bundle", "บทบาทหน้าที่", None, None, "จำนวน", None, None, None, None,
                "รายการ tool_id"], height=22)
    r += 1
    for k, b in C.BUNDLES.items():
        cellfont(ws.cell(row=r, column=1, value=k), bold=True, size=8, color=ORANGE)
        c = cellfont(ws.cell(row=r, column=2, value=b["role_th"]), size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cellfont(ws.cell(row=r, column=5, value=len(b["tools"])), size=8).alignment = \
            Alignment(horizontal="center")
        c = cellfont(ws.cell(row=r, column=10, value=", ".join(b["tools"])), size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    cellfont(ws.cell(row=r + 1, column=1,
                     value="ค่า vCPU/RAM/Disk ข้างบนคำนวณด้วยโหมด realistic ที่ 36 เดือน Scale 1.0× "
                           "เพื่อใช้เป็นจุดตั้งต้น — ค่าที่ต้องยื่นขอจริงให้ดูจากชีท 04 "
                           "ซึ่งใช้โหมดและพารามิเตอร์ที่ตั้งไว้ในชีท 05"),
             size=8, italic=True, color="808080")
    return ws


# =========================================================================== #
def sheet_readme(wb):
    ws = wb.active
    ws.title = S_README
    widths(ws, {"A": 4, "B": 30, "C": 110})
    ws["B2"] = "CI/CD Resource & Compliance Planner"
    ws["B2"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
    ws["B3"] = "ตารางเครื่องมือ CI/CD ทุกประเภท + การคำนวณทรัพยากรกรณีเครื่องแชร์ + ผลลัพธ์ระยะยาว"
    ws["B3"].font = Font(name=FONT, size=11, color=ORANGE)
    ws["B4"] = (f"Schema {C.SCHEMA_VERSION} — ใช้เป็นแบบฟอร์มกลางสำหรับประเมินโครงการใด ๆ "
                "(ไม่มีข้อมูลเฉพาะโครงการอยู่ในไฟล์นี้)")
    ws["B4"].font = Font(name=FONT, size=9, italic=True, color="808080")

    blocks = [
        ("ชีทในไฟล์นี้", [
            (S_PARAM, "พารามิเตอร์ทั้งหมดของโมเดล — เปลี่ยนโหมด strict/realistic, Scale Factor, OS Reserve, "
                      "ช่วงเวลาประเมิน (ช่องสีเหลืองคือช่องที่แก้ได้)"),
            (S_TOOLS, f"ตารางเครื่องมือ {NT} รายการ ครบ 6 Stage พร้อม minimum vCPU/RAM/Disk, ที่มาของตัวเลข, "
                      "มาตรฐานไทย/สากลที่แต่ละเครื่องมือช่วยตอบ, ทางเลือก Enterprise/OSS และข้อควรระวัง"),
            (S_FREQ, "ชั้นความถี่การรัน 7 ระดับ และการแปลงเป็นน้ำหนัก 50%-95%"),
            (S_CALC, "ตารางกำหนดว่าเครื่องมือใดติดตั้งบน VM ใด (ใส่ 1/0) — เป็นช่องกรอกหลัก"),
            (S_VM, "ผลการคำนวณต่อ VM: เงื่อนไข A / B1 / B2 / C, ค่าที่ต้องจัดสรร, และส่วนต่างกับ spec จริง"),
            (S_STORE, "ผลลัพธ์ระยะยาวของพื้นที่จัดเก็บที่ 12 / 24 / 36 / 60 เดือน + สิ่งที่ต้องทำเพื่อคุมการโต"),
            (S_COMP, f"มาตรการที่ต้องทำ {len(C.CONTROLS)} รายการ พร้อมสถานะผ่าน/ไม่ผ่าน "
                     "และเลขข้อของมาตรฐานที่อ้างแต่ละมาตรการ"),
            (S_CAPMTX, "เมทริกซ์เครื่องมือ x capability (ฐานคำนวณของชีท 07)"),
            (S_PROFILE, "เปรียบเทียบประเภทโครงการ 5 แบบ"),
            (S_ARCH, "ผังเครื่องอ้างอิง 4 แบบ (2 / 4 / 6 เครื่อง และ AI/ML) พร้อมชุดเครื่องมือตามบทบาท "
                     "— ใช้เป็นจุดตั้งต้นของโครงการใหม่"),
            (S_FW, f"ทะเบียนกฎหมายและมาตรฐาน {len(C.FRAMEWORKS)} ฉบับ แยกไทย/สากล "
                   "พร้อมเมทริกซ์ว่าฉบับใดอ้างมาตรการใด (เลขข้ออยู่ใน comment ของช่อง)"),
        ]),
        ("วิธีใช้", [
            ("ขั้นที่ 1", f"เปิดชีท {S_PARAM} เลือกโหมดการคำนวณและ Scale Factor ให้ตรงกับขนาดงานจริง"),
            ("ขั้นที่ 2", f"เลือกผังตั้งต้นจากชีท {S_ARCH} แล้วเปิดชีท {S_CALC} "
                          "ใส่ 1 ในคอลัมน์ VM ที่จะติดตั้งเครื่องมือนั้น "
                          "(VM01-06 ใส่ผังอ้างอิงภาครัฐ 6 เครื่องไว้ให้แล้ว, VM07-10 ว่างสำหรับกรณีอื่น)"),
            ("ขั้นที่ 3", f"ดูผลที่ชีท {S_VM} — คอลัมน์ 'จัดสรร' คือค่า minimum ที่ต้องขอ "
                          "และคอลัมน์ 'ผลประเมิน' บอกว่า spec ที่ขอไว้พอหรือไม่"),
            ("ขั้นที่ 4", f"ตรวจ compliance ที่ชีท {S_COMP} — ข้อที่ 'ไม่ผ่าน' คือ capability ที่ยังไม่มีเครื่องมือรองรับ"),
            ("ขั้นที่ 5", f"วางแผนพื้นที่จัดเก็บระยะยาวจากชีท {S_STORE}"),
        ]),
        ("รหัสสีที่ใช้", [
            ("ตัวเลขสีน้ำเงิน + พื้นเหลือง", "ช่องที่ผู้ใช้กรอก/แก้ไขได้"),
            ("ตัวเลขสีเขียว", "สูตรที่อ้างอิงข้ามชีท"),
            ("ตัวเลขสีดำ", "สูตรคำนวณภายในชีทเดียวกัน"),
            ("พื้นเขียวอ่อน", "ผลลัพธ์สุดท้ายที่ต้องนำไปใช้ขอทรัพยากร"),
            ("พื้นสีตาม Stage", "ฟ้า=Stage 1, เหลือง=2, ส้ม=3, เขียว=4, ม่วง=5, พีช=6"),
        ]),
        ("ข้อจำกัดที่ต้องรู้ก่อนใช้ตัวเลขนี้", [
            ("1", "ตัวเลข minimum เป็นค่าตั้งต้นจากเอกสารติดตั้งของผู้พัฒนาแต่ละเครื่องมือ "
                  "รวมกับค่าที่พบจากการใช้งานจริงระดับ UAT/Production ขนาดเล็ก "
                  "ไม่ใช่ผลวัดจากระบบของท่าน — ต้องวัด baseline จริง 2-4 สัปดาห์แล้วปรับ"),
            ("2", "โหมด strict (ตามโจทย์) จะได้ค่าสูงกว่าความเป็นจริงในเครื่องที่มีเครื่องมือ ephemeral จำนวนมาก "
                  "เพราะบวกทุกตัวแม้ในความจริงจะรันเรียงต่อกัน — ใช้ยื่นขอทรัพยากรเพื่อความปลอดภัย "
                  "และใช้โหมด realistic เพื่อดูค่าที่น่าจะเกิดขึ้นจริง"),
            ("3", "ปริมาณข้อมูลต่อวันอ่อนไหวต่อระดับ log และจำนวน build มาก — เปลี่ยน Scale Factor "
                  "ให้ตรงกับปริมาณจริงก่อนอ่านผลชีท 06"),
            ("3.1", "คอลัมน์มาตรฐานในชีท 01 บอกว่าเครื่องมือ 'ช่วยตอบ' ข้อกำหนดใดได้ "
                    "ไม่ได้รับประกันว่าตั้งค่าถูกต้อง — การผ่านมาตรฐานจริงยังต้องมีการตั้งค่า "
                    "กระบวนการ และหลักฐานการตรวจสอบประกอบ"),
            ("4", "ตารางนี้ไม่ครอบคลุม network bandwidth, IOPS และ license cost ซึ่งต้องประเมินแยก"),
            ("5", "License ที่เป็น GPL/AGPL (MinIO, Grafana, Zabbix, FOSSology, Wazuh, testssl.sh) "
                  "ขัดกับข้อห้ามของโครงการภาครัฐบางแห่ง — ต้องตรวจเงื่อนไขการใช้งานก่อน"),
        ]),
    ]
    r = 6
    for head, items in blocks:
        ws.cell(row=r, column=2, value=head).font = Font(name=FONT, size=12, bold=True, color=NAVY)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=GREY)
        r += 1
        for a, b in items:
            cellfont(ws.cell(row=r, column=2, value=a), bold=True, size=9, color=ORANGE)
            c = ws.cell(row=r, column=3, value=b)
            cellfont(c, size=9, color="404040")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1
    ws["C6"].comment = Comment("ไฟล์นี้สร้างจาก scripts/build_xlsx.py ใน repo cicd-resource-planner "
                               "แก้ข้อมูลต้นทางที่ scripts/catalog_data.py แล้ว build ใหม่",
                               "CI/CD Planner")
    return ws


# =========================================================================== #
def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "dist", "CICD_Tool_Resource_Matrix.xlsx")
    wb = Workbook()
    sheet_readme(wb)
    sheet_tools(wb)
    sheet_freq(wb)
    sheet_calc(wb)
    sheet_vm(wb)
    sheet_params(wb)
    sheet_storage(wb)
    _, caps, hl = sheet_capmatrix(wb)
    sheet_compliance(wb, caps, hl)
    sheet_profile(wb)
    sheet_archetypes(wb)
    sheet_frameworks(wb)
    # เรียงชีทตามหมายเลข
    order = [S_README, S_TOOLS, S_FREQ, S_CALC, S_VM, S_PARAM, S_STORE, S_COMP, S_CAPMTX, S_PROFILE, S_ARCH, S_FW]
    wb._sheets = [wb[n] for n in order]
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"[ok] {out}")


if __name__ == "__main__":
    main()
